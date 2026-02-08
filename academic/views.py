"""
Views for academic app (P2, P3, P4 - Academic Data Management)
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q
from django.http import JsonResponse
from .models import AcademicYear, Semester, Course, Section, Enrollment
from accounts.models import User, UserProfile
import openpyxl
from openpyxl import load_workbook
from datetime import date


def is_admin(user):
    """Check if user is admin"""
    return user.is_authenticated and user.is_admin()


def is_teacher(user):
    """Check if user is teacher"""
    return user.is_authenticated and user.is_teacher()


@login_required
@user_passes_test(is_admin)
def course_list(request):
    """
    Process 2: จัดการข้อมูลพื้นฐาน - ดูรายวิชา
    """
    from django.db.models import Count, Q
    from academic.models import Enrollment
    
    courses = Course.objects.filter(is_active=True).prefetch_related('sections', 'sections__teacher', 'sections__enrollments')
    
    # Add teacher and student count info to each course
    courses_with_info = []
    for course in courses:
        # Get all teachers teaching this course (from all sections)
        teachers = []
        for section in course.sections.all():
            if section.teacher and section.teacher not in teachers:
                teachers.append(section.teacher)
        
        # Count total enrolled students across all sections
        total_students = Enrollment.objects.filter(
            section__course=course,
            status='enrolled'
        ).values('student').distinct().count()
        
        courses_with_info.append({
            'course': course,
            'teachers': teachers,
            'total_students': total_students,
            'sections': course.sections.all(),  # Add sections for edit teacher link
        })
    
    context = {
        'courses_with_info': courses_with_info,
    }
    return render(request, 'academic/course_list.html', context)


@login_required
@user_passes_test(is_admin)
def course_detail(request, course_id):
    """
    ดูรายละเอียดรายวิชา
    """
    from django.db.models import Count
    from academic.models import Enrollment
    
    course = get_object_or_404(Course, id=course_id, is_active=True)
    
    # Get all sections for this course
    sections = course.sections.all().select_related('teacher', 'semester').order_by('section_number')
    
    # Get section details with teacher and student count
    sections_with_info = []
    for section in sections:
        # Count enrolled students
        student_count = Enrollment.objects.filter(
            section=section,
            status='enrolled'
        ).count()
        
        sections_with_info.append({
            'section': section,
            'teacher': section.teacher,
            'student_count': student_count,
        })
    
    # Get total students across all sections
    total_students = Enrollment.objects.filter(
        section__course=course,
        status='enrolled'
    ).values('student').distinct().count()
    
    # Get all unique teachers
    teachers = []
    for section in sections:
        if section.teacher and section.teacher not in teachers:
            teachers.append(section.teacher)
    
    context = {
        'course': course,
        'sections_with_info': sections_with_info,
        'total_students': total_students,
        'teachers': teachers,
    }
    return render(request, 'academic/course_detail.html', context)


@login_required
@user_passes_test(is_admin)
def course_delete(request, course_id):
    """
    ลบรายวิชา (Admin only)
    """
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        course_code = course.course_code
        course.is_active = False  # Soft delete
        course.save()
        messages.success(request, f'ลบรายวิชา {course_code} สำเร็จ')
        return redirect('academic:course_list')
    
    context = {
        'course': course,
    }
    return render(request, 'academic/course_delete.html', context)


@login_required
@user_passes_test(is_admin)
def manage_course(request, course_id=None):
    """
    Process 2: จัดการข้อมูลพื้นฐาน - เพิ่ม/แก้ไขรายวิชา
    """
    if course_id:
        course = get_object_or_404(Course, id=course_id)
    else:
        course = None
    
    if request.method == 'POST':
        # Check if this is an AJAX request to add a new teacher
        if request.POST.get('action') == 'add_teacher' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            username = request.POST.get('username', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            employee_id = request.POST.get('employee_id', '').strip()
            department = request.POST.get('department', '').strip()
            
            if not username:
                return JsonResponse({'success': False, 'error': 'กรุณากรอกชื่อผู้ใช้'})
            
            # Check if username already exists
            if User.objects.filter(username=username).exists():
                return JsonResponse({'success': False, 'error': 'ชื่อผู้ใช้นี้มีอยู่แล้ว'})
            
            try:
                # Create new teacher user
                teacher = User.objects.create_user(
                    username=username,
                    email=email if email else f'{username}@example.com',
                    first_name=first_name,
                    last_name=last_name,
                    role='teacher',
                    is_staff=False
                )
                teacher.set_password('changeme123')  # Default password
                teacher.save()
                
                # Create teacher profile
                UserProfile.objects.create(
                    user=teacher,
                    teacher_employee_id=employee_id if employee_id else None,
                    teacher_department=department if department else None
                )
                
                teacher_name = teacher.get_full_name() or teacher.username
                return JsonResponse({
                    'success': True,
                    'teacher_id': teacher.id,
                    'teacher_name': teacher_name
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
        
        course_code = request.POST.get('course_code')
        course_name = request.POST.get('course_name')
        section_number = request.POST.get('section_number', '')
        time_room = request.POST.get('time_room', '')
        
        # Parse time_room to schedule and room
        schedule = ''
        room = ''
        if time_room:
            if ' / ' in time_room:
                parts = time_room.split(' / ', 1)
                schedule = parts[0].strip()
                room = parts[1].strip() if len(parts) > 1 else ''
            else:
                # Try to detect if it's schedule or room
                if any(char.isdigit() for char in time_room):
                    schedule = time_room
                else:
                    room = time_room
        
        if course_code and course_name:
            # Get credit from form
            credit = request.POST.get('credit', '3')
            if not credit:
                credit = '3'
            description = request.POST.get('description', '')
            
            if course:
                course.course_code = course_code
                course.course_name = course_name
                course.credit = int(credit) if credit else 3
                course.description = description
                course.save()
                
                # Handle teacher assignments for all sections
                for section in course.sections.all():
                    teacher_id = request.POST.get(f'teacher_{section.id}')
                    if teacher_id:
                        if teacher_id == 'none':
                            section.teacher = None
                        else:
                            teacher = get_object_or_404(User, id=teacher_id, role='teacher')
                            section.teacher = teacher
                        section.save()
                
                messages.success(request, 'แก้ไขรายวิชาและอาจารย์ผู้สอนสำเร็จ')
            else:
                course = Course.objects.create(
                    course_code=course_code,
                    course_name=course_name,
                    credit=int(credit) if credit else 3,
                    description=description
                )
                messages.success(request, 'เพิ่มรายวิชาสำเร็จ')
                
                # Check if existing section is selected or new section should be created
                section_id = request.POST.get('section_id', '')
                if section_id:
                    # Link existing section to this course
                    try:
                        existing_section = Section.objects.get(id=section_id)
                        existing_section.course = course
                        existing_section.save()
                        
                        # Handle teacher assignment for the linked section
                        teacher_id = request.POST.get(f'teacher_{section_id}')
                        if teacher_id:
                            if teacher_id == 'none':
                                existing_section.teacher = None
                            else:
                                teacher = get_object_or_404(User, id=teacher_id, role='teacher')
                                existing_section.teacher = teacher
                            existing_section.save()
                        
                        messages.success(request, f'เชื่อมโยงกลุ่มเรียน {existing_section.section_number} กับรายวิชานี้สำเร็จ')
                    except Section.DoesNotExist:
                        messages.warning(request, 'ไม่พบกลุ่มเรียนที่เลือก')
                elif section_number:
                    # Create new section if section_number is provided
                    from datetime import datetime
                    from .models import AcademicYear, Semester
                    
                    # Get or create current academic year
                    current_year = datetime.now().year + 543  # Thai Buddhist year
                    academic_year, _ = AcademicYear.objects.get_or_create(
                        year=current_year,
                        defaults={'description': f'ปีการศึกษา {current_year}'}
                    )
                    
                    # Get or create current semester (default to semester 1)
                    semester, _ = Semester.objects.get_or_create(
                        academic_year=academic_year,
                        semester_number=1,
                        defaults={
                            'start_date': datetime.now().date(),
                            'end_date': datetime.now().date(),
                            'is_active': True
                        }
                    )
                    
                    # Create section
                    new_section = Section.objects.create(
                        course=course,
                        semester=semester,
                        section_number=section_number,
                        schedule=schedule if schedule else None,
                        room=room if room else None,
                        capacity=30
                    )
                    
                    # Handle teacher assignment for the new section
                    teacher_id = request.POST.get(f'teacher_new_{section_number}')
                    if teacher_id and teacher_id != 'none':
                        teacher = get_object_or_404(User, id=teacher_id, role='teacher')
                        new_section.teacher = teacher
                        new_section.save()
                    
                    messages.success(request, f'เพิ่มกลุ่มเรียน {section_number} สำเร็จ')
            
            return redirect('academic:course_list')
        else:
            messages.error(request, 'กรุณากรอกข้อมูลให้ครบถ้วน')
    
    # Get sections and teachers for editing
    sections = None
    teachers = None
    existing_sections = []
    
    # Always get teachers list
    try:
        teachers = User.objects.filter(role='teacher').order_by('first_name', 'last_name', 'username')
    except Exception:
        teachers = []
    
    # Get all existing sections for dropdown (both for add and edit)
    try:
        # Get all sections that have a course (filter out None courses)
        # Use filter with course__isnull=False to ensure we only get sections with courses
        sections_queryset = Section.objects.filter(
            course__isnull=False
        ).select_related('course', 'semester', 'semester__academic_year', 'teacher')
        
        # Further filter for active courses if the field exists
        try:
            sections_queryset = sections_queryset.filter(course__is_active=True)
        except Exception:
            # If is_active field doesn't exist, just use all sections with courses
            pass
        
        existing_sections = list(sections_queryset.order_by('course__course_code', 'section_number'))
        
        # Add teacher_id to each section for template
        for section in existing_sections:
            try:
                section.teacher_id_value = section.teacher.id if section.teacher else None
            except Exception:
                section.teacher_id_value = None
    except Exception as e:
        # If there's an error, set to empty list
        existing_sections = []
    
    if course:
        sections = course.sections.all().select_related('teacher', 'semester').order_by('section_number')
    
    context = {
        'course': course,
        'sections': sections,
        'teachers': teachers,
        'existing_sections': existing_sections,
    }
    return render(request, 'academic/course_form.html', context)


@login_required
@user_passes_test(is_admin)
def manage_course_students(request, course_id):
    """
    จัดการนักศึกษาในรายวิชา (แสดงรายชื่อนักศึกษาทั้งหมดที่เรียนในรายวิชานี้)
    """
    course = get_object_or_404(Course, id=course_id, is_active=True)
    
    # Get all enrollments for this course (across all sections)
    # Include both 'pending' and 'enrolled' status to show students uploaded from Excel
    from academic.models import Enrollment
    enrollments = Enrollment.objects.filter(
        section__course=course,
        status__in=['pending', 'enrolled', 'withdrawn', 'completed']
    ).select_related('student', 'section', 'student__profile').order_by('student__first_name', 'student__last_name', 'student__username')
    
    # Get all sections for this course
    sections = course.sections.all().order_by('section_number')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        enrollments = enrollments.filter(
            Q(student__username__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__profile__student_id__icontains=search_query)
        )
    
    # Filter by section if provided
    section_filter = request.GET.get('section_id')
    if section_filter:
        enrollments = enrollments.filter(section_id=section_filter)
    
    # Get all students (for adding new students)
    all_students = User.objects.filter(role='student').order_by('first_name', 'last_name', 'username')
    enrolled_student_ids = enrollments.values_list('student_id', flat=True).distinct()
    available_students = all_students.exclude(id__in=enrolled_student_ids)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            student_id = request.POST.get('student_id')
            section_id = request.POST.get('section_id')
            
            if student_id and section_id:
                student = get_object_or_404(User, id=student_id, role='student')
                section = get_object_or_404(Section, id=section_id, course=course)
                
                enrollment, created = Enrollment.objects.get_or_create(
                    student=student,
                    section=section,
                    defaults={'status': 'enrolled'}
                )
                if created:
                    messages.success(request, f'เพิ่มนักศึกษา {student.get_full_name() or student.username} สำเร็จ')
                else:
                    if enrollment.status != 'enrolled':
                        enrollment.status = 'enrolled'
                        enrollment.save()
                        messages.success(request, f'เพิ่มนักศึกษา {student.get_full_name() or student.username} สำเร็จ')
                    else:
                        messages.info(request, 'นักศึกษาคนนี้ลงทะเบียนอยู่แล้ว')
                return redirect('academic:manage_course_students', course_id=course_id)
        
        elif action == 'edit':
            enrollment_id = request.POST.get('enrollment_id')
            new_section_id = request.POST.get('section_id')
            new_status = request.POST.get('status')
            
            if enrollment_id:
                enrollment = get_object_or_404(Enrollment, id=enrollment_id, section__course=course)
                if new_section_id:
                    new_section = get_object_or_404(Section, id=new_section_id, course=course)
                    enrollment.section = new_section
                if new_status:
                    enrollment.status = new_status
                enrollment.save()
                messages.success(request, 'แก้ไขข้อมูลสำเร็จ')
                return redirect('academic:manage_course_students', course_id=course_id)
        
        elif action == 'remove':
            enrollment_id = request.POST.get('enrollment_id')
            if enrollment_id:
                enrollment = get_object_or_404(Enrollment, id=enrollment_id, section__course=course)
                student_name = enrollment.student.get_full_name() or enrollment.student.username
                enrollment.delete()
                messages.success(request, f'ลบนักศึกษา {student_name} ออกจากรายวิชาสำเร็จ')
                return redirect('academic:manage_course_students', course_id=course_id)
    
    context = {
        'course': course,
        'enrollments': enrollments,
        'sections': sections,
        'available_students': available_students,
        'search_query': search_query,
        'selected_section': section_filter,
    }
    return render(request, 'academic/manage_course_students.html', context)


@login_required
@user_passes_test(is_admin)
def manage_course_teachers(request, course_id):
    """
    จัดการอาจารย์ผู้สอนสำหรับรายวิชา (แก้ไขอาจารย์สำหรับทุก section)
    """
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == 'POST':
        # Handle multiple section teacher assignments
        for section in course.sections.all():
            teacher_id = request.POST.get(f'teacher_{section.id}')
            if teacher_id:
                if teacher_id == 'none':
                    section.teacher = None
                else:
                    teacher = get_object_or_404(User, id=teacher_id, role='teacher')
                    section.teacher = teacher
            section.save()
        
        messages.success(request, f'แก้ไขอาจารย์ผู้สอนสำหรับรายวิชา {course.course_code} สำเร็จ')
        return redirect('academic:course_list')
    
    # Get all sections for this course
    sections = course.sections.all().select_related('teacher', 'semester').order_by('section_number')
    
    # Get all teachers
    teachers = User.objects.filter(role='teacher').order_by('first_name', 'last_name', 'username')
    
    context = {
        'course': course,
        'sections': sections,
        'teachers': teachers,
    }
    return render(request, 'academic/manage_course_teachers.html', context)


@login_required
@user_passes_test(is_admin)
def assign_teacher(request, section_id):
    """
    Process 3: กำหนดผู้สอน - กำหนดอาจารย์ให้กับกลุ่มเรียน
    """
    section = get_object_or_404(Section, id=section_id)
    
    if request.method == 'POST':
        teacher_id = request.POST.get('teacher_id')
        if teacher_id:
            teacher = get_object_or_404(User, id=teacher_id, role='teacher')
            section.teacher = teacher
            section.save()
            messages.success(request, f'กำหนดอาจารย์ {teacher.get_full_name()} ให้กับกลุ่มเรียนสำเร็จ')
            return redirect('academic:section_list')
        else:
            section.teacher = None
            section.save()
            messages.success(request, 'ลบการกำหนดอาจารย์สำเร็จ')
            return redirect('academic:section_list')
    
    teachers = User.objects.filter(role='teacher')
    context = {
        'section': section,
        'teachers': teachers,
    }
    return render(request, 'academic/assign_teacher.html', context)


@login_required
@user_passes_test(is_teacher)
def my_sections(request):
    """
    หน้าสำหรับอาจารย์ดูกลุ่มเรียนที่สอน
    """
    sections = Section.objects.filter(teacher=request.user)
    context = {
        'sections': sections,
    }
    return render(request, 'academic/my_sections.html', context)


@login_required
@user_passes_test(is_teacher)
def manage_enrollment(request, section_id):
    """
    Process 4: บันทึกข้อมูลผู้เรียน - อาจารย์จัดการนักศึกษาในกลุ่มเรียน
    """
    section = get_object_or_404(Section, id=section_id, teacher=request.user)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        student_id = request.POST.get('student_id')
        
        if action == 'add' and student_id:
            student = get_object_or_404(User, id=student_id, role='student')
            enrollment, created = Enrollment.objects.get_or_create(
                student=student,
                section=section,
                defaults={'status': 'enrolled'}
            )
            if created:
                messages.success(request, f'เพิ่มนักศึกษา {student.get_full_name()} สำเร็จ')
            else:
                messages.info(request, 'นักศึกษาคนนี้ลงทะเบียนอยู่แล้ว')
        
        elif action == 'remove' and student_id:
            enrollment = Enrollment.objects.filter(student_id=student_id, section=section).first()
            if enrollment:
                enrollment.delete()
                messages.success(request, 'ลบนักศึกษาออกจากกลุ่มเรียนสำเร็จ')
    
    enrollments = Enrollment.objects.filter(section=section, status='enrolled').select_related('student', 'student__profile')
    all_students = User.objects.filter(role='student').select_related('profile')
    enrolled_student_ids = enrollments.values_list('student_id', flat=True)
    available_students = all_students.exclude(id__in=enrolled_student_ids)
    
    context = {
        'section': section,
        'enrollments': enrollments,
        'available_students': available_students,
    }
    return render(request, 'academic/manage_enrollment.html', context)


@login_required
@user_passes_test(is_teacher)
def import_students_excel(request):
    """
    หน้าเมนูผู้เรียน - อาจารย์โหลด Excel เพื่อเพิ่มข้อมูลผู้เรียน
    """
    # Get all academic years (including past years)
    academic_years = AcademicYear.objects.all().order_by('-year')
    
    # Get courses that this teacher teaches
    teacher_sections = Section.objects.filter(teacher=request.user).select_related('course', 'semester')
    teacher_courses = Course.objects.filter(
        sections__teacher=request.user
    ).distinct().order_by('course_code')
    
    if request.method == 'POST':
        academic_year_id = request.POST.get('academic_year')
        semester_number = request.POST.get('semester_number')  # 1, 2, or 3
        course_id = request.POST.get('course')
        section_id = request.POST.get('section')
        excel_file = request.FILES.get('excel_file')
        
        if not all([academic_year_id, semester_number, course_id, section_id, excel_file]):
            messages.error(request, 'กรุณากรอกข้อมูลให้ครบถ้วน')
        else:
            try:
                # Get section
                section = get_object_or_404(Section, id=section_id, teacher=request.user)
                
                # Load Excel file
                workbook = load_workbook(excel_file, read_only=True)
                worksheet = workbook.active
                
                # Expected columns: รหัสนักศึกษา, ชื่อ, นามสกุล
                # Try to find header row
                header_row = None
                student_id_col = None
                first_name_col = None
                last_name_col = None
                
                for row_idx, row in enumerate(worksheet.iter_rows(max_row=10, values_only=True), start=1):
                    row_lower = [str(cell).lower().strip() if cell else '' for cell in row]
                    if 'รหัส' in ' '.join(row_lower) or 'student' in ' '.join(row_lower):
                        header_row = row_idx
                        # Find column indices
                        for col_idx, cell in enumerate(row, start=1):
                            cell_str = str(cell).lower().strip() if cell else ''
                            if 'รหัส' in cell_str or 'student' in cell_str or 'id' in cell_str:
                                student_id_col = col_idx
                            elif 'ชื่อ' in cell_str and 'นามสกุล' not in cell_str or 'first' in cell_str:
                                first_name_col = col_idx
                            elif 'นามสกุล' in cell_str or 'last' in cell_str or 'surname' in cell_str:
                                last_name_col = col_idx
                        break
                
                # If no header found, assume first row is header and use first 3 columns
                if header_row is None:
                    header_row = 1
                    student_id_col = 1
                    first_name_col = 2
                    last_name_col = 3
                
                # Process data rows
                success_count = 0
                error_count = 0
                errors = []
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    # Get values
                    student_id = str(row[student_id_col - 1]).strip() if student_id_col and len(row) >= student_id_col else None
                    first_name = str(row[first_name_col - 1]).strip() if first_name_col and len(row) >= first_name_col else ''
                    last_name = str(row[last_name_col - 1]).strip() if last_name_col and len(row) >= last_name_col else ''
                    
                    if not student_id or student_id == 'None':
                        continue
                    
                    try:
                        # Check if UserProfile exists with this student_id
                        student_profile = UserProfile.objects.filter(student_id=student_id).first()
                        
                        if student_profile:
                            # User already exists, use it
                            student = student_profile.user
                            # Update name if provided and user doesn't have name
                            if not student.first_name and first_name:
                                student.first_name = first_name
                            if not student.last_name and last_name:
                                student.last_name = last_name
                            student.save()
                        else:
                            # Create new user and student profile
                            # Generate username from student_id
                            username = f"student_{student_id}"
                            # Make sure username is unique
                            base_username = username
                            counter = 1
                            while User.objects.filter(username=username).exists():
                                username = f"{base_username}_{counter}"
                                counter += 1
                            
                            # Create user (name can be empty)
                            student = User.objects.create_user(
                                username=username,
                                email=f"{username}@example.com",  # Temporary email
                                password='changeme123',  # Default password, user should change
                                first_name=first_name if first_name else '',
                                last_name=last_name if last_name else '',
                                role='student'
                            )
                            
                            # Create student profile
                            UserProfile.objects.create(
                                user=student,
                                student_id=student_id
                            )
                        
                        # Create enrollment with pending status
                        enrollment, created = Enrollment.objects.get_or_create(
                            student=student,
                            section=section,
                            defaults={'status': 'pending'}
                        )
                        
                        if created:
                            success_count += 1
                        else:
                            # If enrollment exists but is pending, keep it as pending
                            # If it's already enrolled, skip
                            if enrollment.status == 'pending':
                                pass  # Keep as pending until student registers
                            # Already enrolled, skip
                            pass
                            
                    except Exception as e:
                        error_count += 1
                        errors.append(f"แถว {row_idx}: {str(e)}")
                
                if success_count > 0:
                    messages.success(request, f'โหลดข้อมูลสำเร็จ {success_count} รายการ')
                if error_count > 0:
                    messages.warning(request, f'มีข้อผิดพลาด {error_count} รายการ')
                if errors:
                    for error in errors[:10]:  # Show first 10 errors
                        messages.error(request, error)
                
                return redirect('academic:import_students_excel')
                
            except Exception as e:
                messages.error(request, f'เกิดข้อผิดพลาดในการโหลดไฟล์: {str(e)}')
    
    context = {
        'academic_years': academic_years,
        'teacher_courses': teacher_courses,
    }
    return render(request, 'academic/import_students_excel.html', context)


@login_required
@user_passes_test(is_teacher)
def get_sections_by_course(request):
    """
    API endpoint: ดึงกลุ่มเรียนตามวิชา, ปีการศึกษา, และภาคเรียน
    """
    course_id = request.GET.get('course_id')
    academic_year_id = request.GET.get('academic_year_id')
    semester_number = request.GET.get('semester_number')
    
    if not all([course_id, academic_year_id, semester_number]):
        return JsonResponse({'error': 'กรุณากรอกข้อมูลให้ครบถ้วน'}, status=400)
    
    try:
        # Get semester
        semester = Semester.objects.get(
            academic_year_id=academic_year_id,
            semester_number=int(semester_number)
        )
        
        # Get sections for this course, semester, and teacher
        sections = Section.objects.filter(
            course_id=course_id,
            semester=semester,
            teacher=request.user
        ).order_by('section_number')
        
        sections_data = [{
            'id': section.id,
            'section_number': section.section_number,
            'room': section.room or '',
            'display': f"{section.section_number} ({section.room or 'ไม่มีห้อง'})"
        } for section in sections]
        
        return JsonResponse({'sections': sections_data})
    except Semester.DoesNotExist:
        return JsonResponse({'error': 'ไม่พบภาคเรียน'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def section_list(request):
    """
    หน้าสำหรับอาจารย์โหลด Excel เพื่อบันทึกข้อมูลผู้เรียน
    แบ่งเป็น 2 ส่วน:
    1. ฟอร์มเลือกข้อมูล: ปีการศึกษา, ภาคเรียน, รายวิชา
    2. อัปโหลด Excel: เลือกกลุ่มเรียน และอัปโหลดไฟล์ Excel
    """
    # Get all academic years (including past years)
    academic_years = AcademicYear.objects.all().order_by('-year')
    
    # Get courses - if admin, show all courses; if teacher, show only courses they teach
    if request.user.is_admin():
        teacher_courses = Course.objects.all().distinct().order_by('course_code')
    else:
        teacher_courses = Course.objects.filter(
            sections__teacher=request.user
        ).distinct().order_by('course_code')
    
    # Get selected values from GET or POST
    selected_academic_year_id = request.GET.get('academic_year') or request.POST.get('academic_year')
    selected_semester_number = request.GET.get('semester_number') or request.POST.get('semester_number')
    selected_course_id = request.GET.get('course') or request.POST.get('course')
    selected_teacher_id = request.GET.get('teacher') or request.POST.get('teacher')
    selected_section_id = request.POST.get('section')
    
    # Convert to appropriate types
    if selected_academic_year_id:
        try:
            selected_academic_year_id = int(selected_academic_year_id)
        except (ValueError, TypeError):
            selected_academic_year_id = None
    
    if selected_semester_number:
        try:
            selected_semester_number = int(selected_semester_number)
        except (ValueError, TypeError):
            selected_semester_number = None
    
    if selected_teacher_id:
        try:
            selected_teacher_id = int(selected_teacher_id)
        except (ValueError, TypeError):
            selected_teacher_id = None
    
    # Get all teachers (for admin filter)
    teachers = User.objects.filter(role='teacher').order_by('first_name', 'last_name', 'username')
    
    # Get semesters for selected academic year
    semesters = []
    if selected_academic_year_id:
        academic_year = AcademicYear.objects.filter(id=selected_academic_year_id).first()
        if academic_year:
            semesters = Semester.objects.filter(
                academic_year=academic_year
            ).order_by('semester_number')
    
    # Get courses for selected semester
    courses = []
    if selected_academic_year_id and selected_semester_number:
        semester = Semester.objects.filter(
            academic_year_id=selected_academic_year_id,
            semester_number=selected_semester_number
        ).first()
        if semester:
            if request.user.is_admin():
                # If teacher filter is selected, filter by teacher
                if selected_teacher_id:
                    # Get courses from sections directly - more reliable
                    # First, get all sections for this teacher in this semester
                    sections_for_teacher = Section.objects.filter(
                        semester=semester,
                        teacher_id=selected_teacher_id
                    ).select_related('course', 'teacher')
                    
                    if sections_for_teacher.exists():
                        # Get unique course IDs from these sections
                        course_ids = sections_for_teacher.values_list('course_id', flat=True).distinct()
                        # Get courses from these IDs - convert to list to ensure it's evaluated
                        courses = list(Course.objects.filter(
                            id__in=list(course_ids)
                        ).prefetch_related('sections', 'sections__teacher').order_by('course_code'))
                    else:
                        # No sections found for this teacher in this semester
                        courses = []
                else:
                    # Get all courses that have sections in this semester
                    courses = list(Course.objects.filter(
                        sections__semester=semester
                    ).distinct().prefetch_related('sections', 'sections__teacher').order_by('course_code'))
            else:
                courses = list(Course.objects.filter(
                    sections__teacher=request.user,
                    sections__semester=semester
                ).distinct().prefetch_related('sections', 'sections__teacher').order_by('course_code'))
    
    # Get sections for selected course
    sections = []
    sections_with_info = []  # For displaying sections table
    if selected_course_id and selected_academic_year_id and selected_semester_number:
        semester = Semester.objects.filter(
            academic_year_id=selected_academic_year_id,
            semester_number=selected_semester_number
        ).first()
        if semester:
            if request.user.is_admin():
                # If teacher filter is selected, filter by teacher
                if selected_teacher_id:
                    sections = Section.objects.filter(
                        course_id=selected_course_id,
                        semester=semester,
                        teacher_id=selected_teacher_id
                    ).select_related('teacher', 'semester', 'course').order_by('section_number')
                else:
                    sections = Section.objects.filter(
                        course_id=selected_course_id,
                        semester=semester
                    ).select_related('teacher', 'semester', 'course').order_by('section_number')
            else:
                sections = Section.objects.filter(
                    course_id=selected_course_id,
                    semester=semester,
                    teacher=request.user
                ).select_related('teacher', 'semester', 'course').order_by('section_number')
            
            # Prepare sections_with_info for table display
            from academic.models import Enrollment
            for section in sections:
                # Count enrolled and pending students
                student_count = Enrollment.objects.filter(
                    section=section,
                    status__in=['pending', 'enrolled']
                ).count()
                
                sections_with_info.append({
                    'section': section,
                    'teacher': section.teacher,
                    'student_count': student_count,
                })
    
    # Handle Excel upload
    if request.method == 'POST' and 'excel_file' in request.FILES:
        excel_file = request.FILES.get('excel_file')
        
        if not all([selected_academic_year_id, selected_semester_number, selected_course_id, selected_section_id, excel_file]):
            messages.error(request, 'กรุณากรอกข้อมูลให้ครบถ้วน')
        else:
            try:
                # Get section - admin can access any section, teacher only their own
                if request.user.is_admin():
                    section = get_object_or_404(Section, id=selected_section_id)
                else:
                    section = get_object_or_404(
                        Section, 
                        id=selected_section_id, 
                        teacher=request.user
                    )
                
                # Load Excel file
                workbook = load_workbook(excel_file, read_only=True)
                worksheet = workbook.active
                
                # Find header row
                header_row = None
                student_id_col = None
                
                for row_idx, row in enumerate(worksheet.iter_rows(max_row=10, values_only=True), start=1):
                    row_lower = [str(cell).lower().strip() if cell else '' for cell in row]
                    if 'รหัส' in ' '.join(row_lower) or 'student' in ' '.join(row_lower):
                        header_row = row_idx
                        for col_idx, cell in enumerate(row, start=1):
                            cell_str = str(cell).lower().strip() if cell else ''
                            if 'รหัส' in cell_str or 'student' in cell_str or 'id' in cell_str:
                                student_id_col = col_idx
                                break
                        break
                
                # If no header found, assume first column is student_id
                if header_row is None:
                    header_row = 1
                    student_id_col = 1
                
                # Process data rows
                success_count = 0
                error_count = 0
                errors = []
                
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    # Get student_id
                    student_id = str(row[student_id_col - 1]).strip() if student_id_col and len(row) >= student_id_col else None
                    
                    if not student_id or student_id == 'None':
                        continue
                    
                    try:
                        # Check if UserProfile exists with this student_id
                        student_profile = UserProfile.objects.filter(student_id=student_id).first()
                        
                        if student_profile:
                            student = student_profile.user
                        else:
                            # Create new user and student profile
                            username = f"student_{student_id}"
                            base_username = username
                            counter = 1
                            while User.objects.filter(username=username).exists():
                                username = f"{base_username}_{counter}"
                                counter += 1
                            
                            student = User.objects.create_user(
                                username=username,
                                email=f"{username}@example.com",
                                password='changeme123',
                                role='student'
                            )
                            
                            UserProfile.objects.create(
                                user=student,
                                student_id=student_id
                            )
                        
                        # Create enrollment with pending status
                        enrollment, created = Enrollment.objects.get_or_create(
                            student=student,
                            section=section,
                            defaults={'status': 'pending'}
                        )
                        
                        if created:
                            success_count += 1
                        else:
                            # If enrollment exists but is pending, keep it as pending
                            # If it's already enrolled, skip
                            if enrollment.status == 'pending':
                                pass  # Keep as pending until student registers
                            else:
                                error_count += 1
                                errors.append(f'รหัสนักศึกษา {student_id} ลงทะเบียนอยู่แล้ว')
                    
                    except Exception as e:
                        error_count += 1
                        errors.append(f'รหัสนักศึกษา {student_id}: {str(e)}')
                
                if success_count > 0:
                    messages.success(request, f'บันทึกข้อมูลสำเร็จ {success_count} รายการ')
                if error_count > 0:
                    messages.warning(request, f'มีข้อผิดพลาด {error_count} รายการ')
                if errors:
                    for error in errors[:10]:  # Show first 10 errors
                        messages.error(request, error)
                
            except Exception as e:
                messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
    
    # Debug info (can be removed in production)
    debug_info = {}
    if selected_teacher_id and selected_academic_year_id and selected_semester_number:
        semester_obj = Semester.objects.filter(
            academic_year_id=selected_academic_year_id,
            semester_number=selected_semester_number
        ).first()
        if semester_obj:
            sections_count = Section.objects.filter(
                semester=semester_obj,
                teacher_id=selected_teacher_id
            ).count()
            debug_info['sections_count'] = sections_count
            debug_info['teacher_id'] = selected_teacher_id
            debug_info['semester_id'] = semester_obj.id
            
            # Debug: Check if teacher exists
            teacher_obj = User.objects.filter(id=selected_teacher_id, role='teacher').first()
            if teacher_obj:
                debug_info['teacher_name'] = teacher_obj.get_full_name() or teacher_obj.username
            else:
                debug_info['teacher_name'] = 'ไม่พบอาจารย์'
            
            # Debug: Check courses
            sections_for_debug = Section.objects.filter(
                semester=semester_obj,
                teacher_id=selected_teacher_id
            )
            course_ids_debug = sections_for_debug.values_list('course_id', flat=True).distinct()
            debug_info['course_ids'] = list(course_ids_debug)
            debug_info['courses_count'] = len(course_ids_debug)
            
            # Debug: Check actual courses query result
            if sections_for_debug.exists():
                course_ids_list = list(course_ids_debug)
                courses_debug = Course.objects.filter(id__in=course_ids_list)
                debug_info['courses_found'] = courses_debug.count()
                debug_info['courses_list'] = [f"{c.course_code} - {c.course_name} (ID: {c.id})" for c in courses_debug]
            else:
                debug_info['courses_found'] = 0
                debug_info['courses_list'] = []
    
    context = {
        'academic_years': academic_years,
        'semesters': semesters,
        'courses': courses,
        'sections': sections,
        'sections_with_info': sections_with_info,  # For sections table display
        'teachers': teachers,
        'selected_academic_year_id': selected_academic_year_id,
        'selected_semester_number': selected_semester_number,
        'selected_course_id': selected_course_id,
        'selected_teacher_id': selected_teacher_id,
        'selected_section_id': selected_section_id,
        'debug_info': debug_info,  # For debugging
    }
    return render(request, 'academic/section_list.html', context)


@login_required
@user_passes_test(is_admin)
def section_add(request):
    """
    เพิ่มกลุ่มเรียนใหม่ (Admin only)
    """
    if request.method == 'POST':
        course_id = request.POST.get('course')
        semester_id = request.POST.get('semester')
        section_number = request.POST.get('section_number')
        teacher_id = request.POST.get('teacher', '')
        capacity = request.POST.get('capacity', 30)
        room = request.POST.get('room', '')
        schedule = request.POST.get('schedule', '')
        
        if course_id and semester_id and section_number:
            course = get_object_or_404(Course, id=course_id)
            semester = get_object_or_404(Semester, id=semester_id)
            
            # Check if section already exists
            if Section.objects.filter(course=course, semester=semester, section_number=section_number).exists():
                messages.error(request, 'กลุ่มเรียนนี้มีอยู่แล้ว')
            else:
                section = Section.objects.create(
                    course=course,
                    semester=semester,
                    section_number=section_number,
                    teacher_id=int(teacher_id) if teacher_id else None,
                    capacity=int(capacity),
                    room=room if room else None,
                    schedule=schedule if schedule else None
                )
                messages.success(request, 'เพิ่มกลุ่มเรียนสำเร็จ')
                return redirect('academic:section_list')
        else:
            messages.error(request, 'กรุณากรอกข้อมูลให้ครบถ้วน')
    
    courses = Course.objects.filter(is_active=True)
    semesters = Semester.objects.filter(is_active=True)
    teachers = User.objects.filter(role='teacher')
    
    context = {
        'courses': courses,
        'semesters': semesters,
        'teachers': teachers,
    }
    return render(request, 'academic/section_form.html', context)


@login_required
@user_passes_test(is_admin)
def section_edit(request, section_id):
    """
    แก้ไขกลุ่มเรียน (Admin only)
    """
    section = get_object_or_404(Section, id=section_id)
    
    if request.method == 'POST':
        course_id = request.POST.get('course')
        semester_id = request.POST.get('semester')
        section_number = request.POST.get('section_number')
        teacher_id = request.POST.get('teacher', '')
        capacity = request.POST.get('capacity', 30)
        room = request.POST.get('room', '')
        schedule = request.POST.get('schedule', '')
        
        if course_id and semester_id and section_number:
            course = get_object_or_404(Course, id=course_id)
            semester = get_object_or_404(Semester, id=semester_id)
            
            # Check if section already exists (excluding current section)
            existing = Section.objects.filter(
                course=course,
                semester=semester,
                section_number=section_number
            ).exclude(id=section_id).exists()
            
            if existing:
                messages.error(request, 'กลุ่มเรียนนี้มีอยู่แล้ว')
            else:
                section.course = course
                section.semester = semester
                section.section_number = section_number
                section.teacher_id = int(teacher_id) if teacher_id else None
                section.capacity = int(capacity)
                section.room = room if room else None
                section.schedule = schedule if schedule else None
                section.save()
                messages.success(request, 'แก้ไขกลุ่มเรียนสำเร็จ')
                return redirect('academic:section_list')
        else:
            messages.error(request, 'กรุณากรอกข้อมูลให้ครบถ้วน')
    
    courses = Course.objects.filter(is_active=True)
    semesters = Semester.objects.filter(is_active=True)
    teachers = User.objects.filter(role='teacher')
    
    context = {
        'section': section,
        'courses': courses,
        'semesters': semesters,
        'teachers': teachers,
    }
    return render(request, 'academic/section_form.html', context)


@login_required
@user_passes_test(is_admin)
def section_delete(request, section_id):
    """
    ลบกลุ่มเรียน (Admin only)
    """
    section = get_object_or_404(Section, id=section_id)
    
    if request.method == 'POST':
        section_name = str(section)
        section.delete()
        messages.success(request, f'ลบกลุ่มเรียน {section_name} สำเร็จ')
        return redirect('academic:section_list')
    
    context = {
        'section': section,
    }
    return render(request, 'academic/section_delete.html', context)


@login_required
@user_passes_test(is_admin)
def create_courses_for_teacher(request):
    """
    สร้างรายวิชาและกลุ่มเรียนสำหรับอาจารย์ thanarat sosawat (ID=2)
    ใช้ปีการศึกษาและภาคเรียนจาก URL parameter หรือใช้ค่าที่เลือกไว้
    """
    if request.method == 'POST':
        try:
            # 1. รับปีการศึกษาและภาคเรียนจาก POST หรือ GET
            academic_year_id = request.POST.get('academic_year') or request.GET.get('academic_year')
            semester_number = request.POST.get('semester_number') or request.GET.get('semester_number')
            teacher_id = request.POST.get('teacher_id') or request.GET.get('teacher', '2')
            
            # 2. ตรวจสอบอาจารย์
            teacher = User.objects.filter(id=teacher_id).first()
            if not teacher:
                teacher = User.objects.filter(username__icontains='thanarat').first()
                if not teacher:
                    teacher = User.objects.filter(role='teacher').first()
                    if not teacher:
                        messages.error(request, 'ไม่พบอาจารย์ในระบบ')
                        return redirect('academic:section_list')
            
            if not academic_year_id or not semester_number:
                # ถ้าไม่มี ให้ใช้ปีการศึกษา 2568 และภาคเรียนที่ 1
                academic_year = AcademicYear.objects.filter(year='2568').first()
                if not academic_year:
                    academic_year = AcademicYear.objects.create(
                        year='2568',
                        description='ปีการศึกษา 2568',
                        is_active=True
                    )
                semester = Semester.objects.filter(
                    academic_year=academic_year,
                    semester_number=1
                ).first()
                if not semester:
                    semester = Semester.objects.create(
                        academic_year=academic_year,
                        semester_number=1,
                        start_date=date(2025, 6, 1),
                        end_date=date(2025, 9, 30),
                        is_active=True
                    )
            else:
                # ใช้ปีการศึกษาและภาคเรียนที่เลือก
                academic_year = AcademicYear.objects.filter(id=academic_year_id).first()
                if not academic_year:
                    messages.error(request, 'ไม่พบปีการศึกษาที่เลือก')
                    return redirect('academic:section_list')
                
                semester = Semester.objects.filter(
                    academic_year=academic_year,
                    semester_number=semester_number
                ).first()
                if not semester:
                    messages.error(request, 'ไม่พบภาคเรียนที่เลือก')
                    return redirect('academic:section_list')
            
            # 4. สร้างรายวิชาใหม่
            courses_data = [
                {
                    'code': 'CS101',
                    'name': 'การเขียนโปรแกรมคอมพิวเตอร์',
                    'name_en': 'Computer Programming',
                    'credit': 3
                },
                {
                    'code': 'CS201',
                    'name': 'โครงสร้างข้อมูลและอัลกอริทึม',
                    'name_en': 'Data Structures and Algorithms',
                    'credit': 3
                },
                {
                    'code': 'CS301',
                    'name': 'ระบบฐานข้อมูล',
                    'name_en': 'Database Systems',
                    'credit': 3
                },
                {
                    'code': 'CS302',
                    'name': 'การพัฒนาเว็บแอปพลิเคชัน',
                    'name_en': 'Web Application Development',
                    'credit': 3
                },
                {
                    'code': 'CS401',
                    'name': 'วิศวกรรมซอฟต์แวร์',
                    'name_en': 'Software Engineering',
                    'credit': 3
                },
            ]
            
            created_count = 0
            section_count = 0
            
            for idx, course_data in enumerate(courses_data, start=1):
                course, created = Course.objects.get_or_create(
                    course_code=course_data['code'],
                    defaults={
                        'course_name': f"{course_data['name']} ({course_data['name_en']})",
                        'credit': course_data['credit'],
                        'is_active': True
                    }
                )
                if created:
                    created_count += 1
                
                # สร้าง Section
                existing_section = Section.objects.filter(
                    course=course,
                    semester=semester,
                    section_number=str(idx)
                ).first()
                
                if existing_section:
                    if existing_section.teacher != teacher:
                        existing_section.teacher = teacher
                        existing_section.save()
                        section_count += 1
                else:
                    Section.objects.create(
                        course=course,
                        semester=semester,
                        section_number=str(idx),
                        teacher=teacher,
                        capacity=30,
                        room=f'LAB-{idx}01',
                        schedule='จันทร์ 09:00-12:00'
                    )
                    section_count += 1
            
            # สร้าง redirect URL พร้อม parameter
            from django.urls import reverse
            from urllib.parse import urlencode
            params = {}
            if academic_year_id:
                params['academic_year'] = academic_year_id
            if semester_number:
                params['semester_number'] = semester_number
            if teacher_id:
                params['teacher'] = teacher_id
            
            redirect_url = reverse('academic:section_list')
            if params:
                redirect_url = f"{redirect_url}?{urlencode(params)}"
            
            messages.success(
                request,
                f'สร้างรายวิชา {created_count} วิชา และกลุ่มเรียน {section_count} กลุ่มสำหรับอาจารย์ {teacher.get_full_name() or teacher.username} สำเร็จ!'
            )
            return redirect(redirect_url)
            
        except Exception as e:
            messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
            # Redirect back with parameters
            from django.urls import reverse
            from urllib.parse import urlencode
            params = {}
            if request.POST.get('academic_year'):
                params['academic_year'] = request.POST.get('academic_year')
            if request.POST.get('semester_number'):
                params['semester_number'] = request.POST.get('semester_number')
            if request.POST.get('teacher_id'):
                params['teacher'] = request.POST.get('teacher_id')
            
            redirect_url = reverse('academic:section_list')
            if params:
                redirect_url = f"{redirect_url}?{urlencode(params)}"
            return redirect(redirect_url)
    
    # GET request - redirect to section_list
    return redirect('academic:section_list')


@login_required
@user_passes_test(is_admin)
def add_students_from_list(request):
    """
    เพิ่มนักเรียน 30 คนจากรายชื่อที่ให้มา
    """
    if request.method == 'POST':
        try:
            # ข้อมูลนักเรียน 30 คน (จาก Excel 28 คน + เพิ่มอีก 2 คน)
            students_data = [
                # 12 คนแรก (เพิ่มแล้ว)
                {'student_id': '68342110008-1', 'fullname': 'นางสาว จรรยพร ทิดี', 'fullname_en': 'Miss JANYAPON THIDEE'},
                {'student_id': '68342110014-9', 'fullname': 'นาย กิตติพัฒน์ โสระมรรค', 'fullname_en': 'Mr. KITTIPAT SORAMAK'},
                {'student_id': '68342110021-7', 'fullname': 'นางสาว ชนัญธิดา ทองค่า', 'fullname_en': 'Miss CHANANTHIDA TONGKHAM'},
                {'student_id': '68342110028-5', 'fullname': 'นาย พนารักษ์ พงษ์เกษม', 'fullname_en': 'Mr. PANARUK PONGKASAM'},
                {'student_id': '68342110035-3', 'fullname': 'นางสาว รัตนา ใจดี', 'fullname_en': 'Miss RATTANA JAIDEE'},
                {'student_id': '68342110042-1', 'fullname': 'นาย สุทธิพงษ์ ขยันเรียน', 'fullname_en': 'Mr. SUTTIPHONG KHAYANRIAN'},
                {'student_id': '68342110049-9', 'fullname': 'นางสาว อรทัย ตั้งใจ', 'fullname_en': 'Miss ORATHAI TANGJAI'},
                {'student_id': '68342110056-7', 'fullname': 'นาย วิทยา มานะ', 'fullname_en': 'Mr. WITTHAYA MANA'},
                {'student_id': '68342110063-5', 'fullname': 'นางสาว กมลชนก ดีใจ', 'fullname_en': 'Miss KAMONCHANOK DEEJAI'},
                {'student_id': '68342110070-3', 'fullname': 'นาย ธีรพงษ์ ขยัน', 'fullname_en': 'Mr. TEERAPHONG KHAYAN'},
                {'student_id': '68342110077-1', 'fullname': 'นางสาว สุดา เรียนดี', 'fullname_en': 'Miss SUDA RIANDEE'},
                # 18 คนที่เหลือ
                {'student_id': '68342110084-9', 'fullname': 'นาย ประเสริฐ ดีงาม', 'fullname_en': 'Mr. PRASERT DEENGAM'},
                {'student_id': '68342110091-7', 'fullname': 'นางสาว มานี ขยันเรียน', 'fullname_en': 'Miss MANEE KHAYANRIAN'},
                {'student_id': '68342110098-5', 'fullname': 'นาย วิชัย เก่งมาก', 'fullname_en': 'Mr. WICHAI KENGMAK'},
                {'student_id': '68342110105-3', 'fullname': 'นางสาว สมหญิง รักเรียน', 'fullname_en': 'Miss SOMYING RAKRIAN'},
                {'student_id': '68342110112-1', 'fullname': 'นาย สมชาย ใจดี', 'fullname_en': 'Mr. SOMCHAI JAIDEE'},
                {'student_id': '68342110119-9', 'fullname': 'นางสาว กาญจนา ตั้งใจ', 'fullname_en': 'Miss KANCHANA TANGJAI'},
                {'student_id': '68342110126-7', 'fullname': 'นาย วรวิทย์ ขยัน', 'fullname_en': 'Mr. WORAWIT KHAYAN'},
                {'student_id': '68342110133-5', 'fullname': 'นางสาว สุภาพ ดีใจ', 'fullname_en': 'Miss SUPHAP DEEJAI'},
                {'student_id': '68342110140-3', 'fullname': 'นาย ธีระ เรียนดี', 'fullname_en': 'Mr. TEERA RIANDEE'},
                {'student_id': '68342110147-1', 'fullname': 'นางสาว อรุณ ดีงาม', 'fullname_en': 'Miss ARUN DEENGAM'},
                {'student_id': '68342110154-9', 'fullname': 'นาย พงศ์ศักดิ์ ขยันเรียน', 'fullname_en': 'Mr. PHONGSAG KHAYANRIAN'},
                {'student_id': '68342110161-7', 'fullname': 'นางสาว วรรณา เก่งมาก', 'fullname_en': 'Miss WANNA KENGMAK'},
                {'student_id': '68342110168-5', 'fullname': 'นาย สุเมธ รักเรียน', 'fullname_en': 'Mr. SUMET RAKRIAN'},
                {'student_id': '68342110175-3', 'fullname': 'นางสาว ปิยะ ใจดี', 'fullname_en': 'Miss PIYA JAIDEE'},
                {'student_id': '68342110182-1', 'fullname': 'นาย ธีรศักดิ์ ตั้งใจ', 'fullname_en': 'Mr. TEERASAK TANGJAI'},
                {'student_id': '68342110189-9', 'fullname': 'นางสาว อรทัย มานะ', 'fullname_en': 'Miss ORATHAI MANA'},
                {'student_id': '68342110196-7', 'fullname': 'นาย วิทยา ดีใจ', 'fullname_en': 'Mr. WITTHAYA DEEJAI'},
                {'student_id': '68342110203-5', 'fullname': 'นางสาว กมลชนก เรียนดี', 'fullname_en': 'Miss KAMONCHANOK RIANDEE'},
                {'student_id': '68342110210-3', 'fullname': 'นาย ธีรพงษ์ ดีงาม', 'fullname_en': 'Mr. TEERAPHONG DEENGAM'},
            ]
            
            def parse_thai_name(fullname):
                """Parse Thai name to extract first name and last name"""
                if fullname.startswith('นางสาว'):
                    name = fullname.replace('นางสาว', '').strip()
                elif fullname.startswith('นาย'):
                    name = fullname.replace('นาย', '').strip()
                else:
                    name = fullname.strip()
                
                parts = name.split()
                if len(parts) >= 2:
                    return parts[0], ' '.join(parts[1:])
                elif len(parts) == 1:
                    return parts[0], ''
                else:
                    return name, ''
            
            created_count = 0
            updated_count = 0
            skipped_count = 0
            
            # รายชื่อที่เพิ่มแล้ว (12 คนแรก)
            already_added = [
                '68342110008-1', '68342110014-9', '68342110021-7', '68342110028-5',
                '68342110035-3', '68342110042-1', '68342110049-9', '68342110056-7',
                '68342110063-5', '68342110070-3', '68342110077-1'
            ]
            
            for student_data in students_data:
                student_id = student_data['student_id']
                
                # ข้ามรายชื่อที่เพิ่มแล้ว
                if student_id in already_added:
                    skipped_count += 1
                    continue
                student_id = student_data['student_id']
                fullname = student_data['fullname']
                
                # Parse Thai name
                first_name, last_name = parse_thai_name(fullname)
                
                # Generate username from student_id (remove hyphen)
                username = student_id.replace('-', '')
                
                # Check if user already exists
                user, created = User.objects.get_or_create(
                    username=username,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'role': 'student',
                        'is_staff': False,
                        'email': f'{username}@student.example.com'
                    }
                )
                
                if created:
                    user.set_password('changeme123')
                    user.save()
                    created_count += 1
                else:
                    # Update existing user
                    user.first_name = first_name
                    user.last_name = last_name
                    user.role = 'student'
                    user.save()
                    updated_count += 1
                
                # Create or update profile
                profile, _ = UserProfile.objects.get_or_create(
                    user=user,
                    defaults={
                        'student_id': student_id
                    }
                )
                
                if profile.student_id != student_id:
                    profile.student_id = student_id
                    profile.save()
            
            messages.success(
                request,
                f'เพิ่มนักเรียนสำเร็จ! สร้างใหม่ {created_count} คน, อัปเดต {updated_count} คน, ข้าม {skipped_count} คน (รวม {created_count + updated_count} คน)'
            )
            return redirect('academic:section_list')
            
        except Exception as e:
            messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
            return redirect('academic:section_list')
    
    # GET request - redirect to section_list
    return redirect('academic:section_list')


@login_required
@user_passes_test(is_admin)
def add_students_to_bis3r1(request):
    """
    เพิ่มนักเรียนเข้าในกลุ่มเรียน bis3r1
    """
    if request.method == 'POST':
        try:
            # ค้นหากลุ่มเรียน bis3r1
            section = Section.objects.filter(section_number__iexact='bis3r1').first()
            
            if not section:
                messages.error(request, 'ไม่พบกลุ่มเรียน bis3r1 ในระบบ')
                return redirect('academic:section_list')
            
            # ตรวจสอบนักเรียนที่มีอยู่
            existing_students = User.objects.filter(role='student')
            
            # สร้างข้อมูลนักเรียนตัวอย่าง (ถ้ายังไม่มี)
            if existing_students.count() < 10:
                students_data = [
                    {'username': 'std001', 'student_id': '68342110001-1', 'first_name': 'สมชาย', 'last_name': 'ใจดี'},
                    {'username': 'std002', 'student_id': '68342110002-2', 'first_name': 'สมหญิง', 'last_name': 'รักเรียน'},
                    {'username': 'std003', 'student_id': '68342110003-3', 'first_name': 'วิชัย', 'last_name': 'เก่งมาก'},
                    {'username': 'std004', 'student_id': '68342110004-4', 'first_name': 'มานี', 'last_name': 'ขยันเรียน'},
                    {'username': 'std005', 'student_id': '68342110005-5', 'first_name': 'ประเสริฐ', 'last_name': 'ดีงาม'},
                    {'username': 'std006', 'student_id': '68342110006-6', 'first_name': 'สุดา', 'last_name': 'เรียนดี'},
                    {'username': 'std007', 'student_id': '68342110007-7', 'first_name': 'ธีรพงษ์', 'last_name': 'มานะ'},
                    {'username': 'std008', 'student_id': '68342110008-8', 'first_name': 'กาญจนา', 'last_name': 'ตั้งใจ'},
                    {'username': 'std009', 'student_id': '68342110009-9', 'first_name': 'วรวิทย์', 'last_name': 'ขยัน'},
                    {'username': 'std010', 'student_id': '68342110010-0', 'first_name': 'สุภาพ', 'last_name': 'ดีใจ'},
                ]
                
                created_count = 0
                for student_data in students_data:
                    student, created = User.objects.get_or_create(
                        username=student_data['username'],
                        defaults={
                            'first_name': student_data['first_name'],
                            'last_name': student_data['last_name'],
                            'role': 'student',
                            'is_staff': False
                        }
                    )
                    if created:
                        student.set_password('changeme123')
                        student.save()
                        
                        # Create profile
                        profile, _ = UserProfile.objects.get_or_create(
                            user=student,
                            defaults={
                                'student_id': student_data['student_id']
                            }
                        )
                        created_count += 1
            
            # เพิ่มนักเรียนเข้าในกลุ่มเรียน bis3r1
            all_students = User.objects.filter(role='student')
            enrolled_students = Enrollment.objects.filter(section=section).values_list('student_id', flat=True)
            available_students = all_students.exclude(id__in=enrolled_students)
            
            # เพิ่มนักเรียน 15 คนแรก (หรือทั้งหมดถ้าน้อยกว่า 15)
            students_to_add = available_students[:15]
            added_count = 0
            
            for student in students_to_add:
                enrollment, created = Enrollment.objects.get_or_create(
                    student=student,
                    section=section,
                    defaults={'status': 'enrolled'}
                )
                if created:
                    added_count += 1
            
            messages.success(
                request,
                f'เพิ่มนักเรียน {added_count} คนเข้าในกลุ่มเรียน {section.course.course_code} - กลุ่ม {section.section_number} สำเร็จ!'
            )
            return redirect('academic:section_list')
            
        except Exception as e:
            messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
            return redirect('academic:section_list')
    
    # GET request - redirect to section_list
    return redirect('academic:section_list')


@login_required
@user_passes_test(is_admin)
def section_detail(request, section_id):
    """
    ดูรายละเอียดกลุ่มเรียน - แสดงข้อมูลกลุ่มเรียนและรายชื่อนักเรียน
    รองรับการค้นหาและกรองข้อมูล
    """
    section = get_object_or_404(Section, id=section_id)
    
    # ดึงข้อมูลการลงทะเบียนทั้งหมดในกลุ่มเรียนนี้
    enrollments = Enrollment.objects.filter(
        section=section
    ).select_related('student', 'student__profile')
    
    # รับค่าการค้นหาและกรองจาก query parameters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    
    # กรองตามสถานะ
    if status_filter:
        enrollments = enrollments.filter(status=status_filter)
    
    # ค้นหาตามคำค้นหา (ชื่อ, นามสกุล, รหัสนักศึกษา, username)
    if search_query:
        enrollments = enrollments.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__username__icontains=search_query) |
            Q(student__profile__student_id__icontains=search_query)
        )
    
    # เรียงลำดับ
    enrollments = enrollments.order_by('student__first_name', 'student__last_name', 'student__username')
    
    # แยกตามสถานะ (สำหรับสถิติ - ใช้ข้อมูลทั้งหมดไม่ใช่แค่ที่กรองแล้ว)
    all_enrollments = Enrollment.objects.filter(section=section)
    pending_students = all_enrollments.filter(status='pending')
    enrolled_students = all_enrollments.filter(status='enrolled')
    withdrawn_students = all_enrollments.filter(status='withdrawn')
    completed_students = all_enrollments.filter(status='completed')
    
    # สถิติ
    total_pending = pending_students.count()
    total_enrolled = enrolled_students.count()
    total_withdrawn = withdrawn_students.count()
    total_completed = completed_students.count()
    total_students = all_enrollments.count()
    
    context = {
        'section': section,
        'enrollments': enrollments,
        'pending_students': pending_students,
        'enrolled_students': enrolled_students,
        'withdrawn_students': withdrawn_students,
        'completed_students': completed_students,
        'total_pending': total_pending,
        'total_enrolled': total_enrolled,
        'total_withdrawn': total_withdrawn,
        'total_completed': total_completed,
        'total_students': total_students,
        'capacity': section.capacity,
        'available_slots': section.capacity - total_enrolled,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    
    return render(request, 'academic/section_detail.html', context)


@login_required
def delete_enrollment(request, section_id, enrollment_id):
    """
    ลบการลงทะเบียนของนักเรียนออกจากกลุ่มเรียน
    """
    section = get_object_or_404(Section, id=section_id)
    enrollment = get_object_or_404(Enrollment, id=enrollment_id, section=section)
    
    # ตรวจสอบสิทธิ์
    if not request.user.is_admin():
        if not (request.user.is_teacher() and section.teacher == request.user):
            messages.error(request, 'คุณไม่มีสิทธิ์ในการลบข้อมูลนี้')
            return redirect('academic:section_detail', section_id=section_id)
    
    if request.method == 'POST':
        student_name = enrollment.student.get_full_name() or enrollment.student.username
        enrollment.delete()
        messages.success(request, f'ลบนักเรียน {student_name} ออกจากกลุ่มเรียนสำเร็จ')
        return redirect('academic:section_detail', section_id=section_id)
    
    # ถ้าไม่ใช่ POST ให้ redirect กลับ
    return redirect('academic:section_detail', section_id=section_id)


@login_required
def batch_delete_enrollments(request, section_id):
    """
    ลบการลงทะเบียนหลายรายการพร้อมกัน
    """
    section = get_object_or_404(Section, id=section_id)
    
    # ตรวจสอบสิทธิ์
    if not request.user.is_admin():
        if not (request.user.is_teacher() and section.teacher == request.user):
            messages.error(request, 'คุณไม่มีสิทธิ์ในการลบข้อมูลนี้')
            return redirect('academic:section_detail', section_id=section_id)
    
    if request.method == 'POST':
        enrollment_ids = request.POST.getlist('enrollment_ids')
        
        if not enrollment_ids:
            messages.warning(request, 'กรุณาเลือกรายการที่ต้องการลบ')
            return redirect('academic:section_detail', section_id=section_id)
        
        # ดึงข้อมูล enrollments ที่เลือก
        enrollments = Enrollment.objects.filter(
            id__in=enrollment_ids,
            section=section
        )
        
        count = enrollments.count()
        
        if count == 0:
            messages.warning(request, 'ไม่พบรายการที่เลือก')
            return redirect('academic:section_detail', section_id=section_id)
        
        # ลบ enrollments
        enrollments.delete()
        
        messages.success(request, f'ลบ {count} รายการสำเร็จ')
        return redirect('academic:section_detail', section_id=section_id)
    
    # ถ้าไม่ใช่ POST ให้ redirect กลับ
    return redirect('academic:section_detail', section_id=section_id)


@login_required
@user_passes_test(is_admin)
def add_students_to_bis3r2(request):
    """
    เพิ่มนักเรียนทั้งหมดจากรายชื่อใหม่เข้าในกลุ่มเรียน bis3r1 (รายวิชา 123456789)
    """
    if request.method == 'POST':
        try:
            # ค้นหากลุ่มเรียน bis3r1 ในรายวิชา 123456789
            section = Section.objects.filter(
                section_number__iexact='bis3r1',
                course__course_code='123456789'
            ).first()
            
            if not section:
                messages.error(request, 'ไม่พบกลุ่มเรียน bis3r1 (รายวิชา 123456789) ในระบบ')
                return redirect('academic:section_list')
            
            # ข้อมูลนักเรียน 27 คนจากรายชื่อใหม่
            students_data = [
                {'student_id': '68342110008-1', 'fullname': 'นางสาว จรรยพร ทิดี', 'fullname_en': 'Miss JANYAPON THIDEE'},
                {'student_id': '68342110014-9', 'fullname': 'นาย กิตติพัฒน์ โสระมรรค', 'fullname_en': 'Mr. KITTIPAT SORAMAK'},
                {'student_id': '68342110021-7', 'fullname': 'นางสาว ชนัญธิดา ทองค่า', 'fullname_en': 'Miss CHANANTHIDA TONGKHAM'},
                {'student_id': '68342110028-5', 'fullname': 'นาย พนารักษ์ พงษ์เกษม', 'fullname_en': 'Mr. PANARUK PONGKASAM'},
                {'student_id': '68342110035-3', 'fullname': 'นางสาว รัตนา ใจดี', 'fullname_en': 'Miss RATTANA JAIDEE'},
                {'student_id': '68342110042-1', 'fullname': 'นาย สุทธิพงษ์ ขยันเรียน', 'fullname_en': 'Mr. SUTTIPHONG KHAYANRIAN'},
                {'student_id': '68342110049-9', 'fullname': 'นางสาว อรทัย ตั้งใจ', 'fullname_en': 'Miss ORATHAI TANGJAI'},
                {'student_id': '68342110056-7', 'fullname': 'นาย วิทยา มานะ', 'fullname_en': 'Mr. WITTHAYA MANA'},
                {'student_id': '68342110063-5', 'fullname': 'นางสาว กมลชนก ดีใจ', 'fullname_en': 'Miss KAMONCHANOK DEEJAI'},
                {'student_id': '68342110070-3', 'fullname': 'นาย ธีรพงษ์ ขยัน', 'fullname_en': 'Mr. TEERAPHONG KHAYAN'},
                {'student_id': '68342110077-1', 'fullname': 'นางสาว สุดา เรียนดี', 'fullname_en': 'Miss SUDA RIANDEE'},
                {'student_id': '68342110084-9', 'fullname': 'นาย ประเสริฐ ดีงาม', 'fullname_en': 'Mr. PRASERT DEENGAM'},
                {'student_id': '68342110091-7', 'fullname': 'นางสาว มานี ขยันเรียน', 'fullname_en': 'Miss MANEE KHAYANRIAN'},
                {'student_id': '68342110098-5', 'fullname': 'นาย วิชัย เก่งมาก', 'fullname_en': 'Mr. WICHAI KENGMAK'},
                {'student_id': '68342110105-3', 'fullname': 'นางสาว สมหญิง รักเรียน', 'fullname_en': 'Miss SOMYING RAKRIAN'},
                {'student_id': '68342110112-1', 'fullname': 'นาย สมชาย ใจดี', 'fullname_en': 'Mr. SOMCHAI JAIDEE'},
                {'student_id': '68342110119-9', 'fullname': 'นางสาว กาญจนา ตั้งใจ', 'fullname_en': 'Miss KANCHANA TANGJAI'},
                {'student_id': '68342110126-7', 'fullname': 'นาย วรวิทย์ ขยัน', 'fullname_en': 'Mr. WORAWIT KHAYAN'},
                {'student_id': '68342110133-5', 'fullname': 'นางสาว สุภาพ ดีใจ', 'fullname_en': 'Miss SUPHAP DEEJAI'},
                {'student_id': '68342110140-3', 'fullname': 'นาย ธีระ เรียนดี', 'fullname_en': 'Mr. TEERA RIANDEE'},
                {'student_id': '68342110147-1', 'fullname': 'นางสาว อรุณ ดีงาม', 'fullname_en': 'Miss ARUN DEENGAM'},
                {'student_id': '68342110154-9', 'fullname': 'นาย พงศ์ศักดิ์ ขยันเรียน', 'fullname_en': 'Mr. PHONGSAG KHAYANRIAN'},
                {'student_id': '68342110161-7', 'fullname': 'นางสาว วรรณา เก่งมาก', 'fullname_en': 'Miss WANNA KENGMAK'},
                {'student_id': '68342110168-5', 'fullname': 'นาย สุเมธ รักเรียน', 'fullname_en': 'Mr. SUMET RAKRIAN'},
                {'student_id': '68342110175-3', 'fullname': 'นางสาว ปิยะ ใจดี', 'fullname_en': 'Miss PIYA JAIDEE'},
                {'student_id': '68342110182-1', 'fullname': 'นาย ธีรศักดิ์ ตั้งใจ', 'fullname_en': 'Mr. TEERASAK TANGJAI'},
                {'student_id': '68342110189-9', 'fullname': 'นางสาว อรทัย มานะ', 'fullname_en': 'Miss ORATHAI MANA'},
                {'student_id': '68342110196-7', 'fullname': 'นาย วิทยา ดีใจ', 'fullname_en': 'Mr. WITTHAYA DEEJAI'},
                {'student_id': '68342110203-5', 'fullname': 'นางสาว กมลชนก เรียนดี', 'fullname_en': 'Miss KAMONCHANOK RIANDEE'},
                {'student_id': '68342110210-3', 'fullname': 'นาย ธีรพงษ์ ดีงาม', 'fullname_en': 'Mr. TEERAPHONG DEENGAM'},
            ]
            
            def parse_thai_name(fullname):
                """Parse Thai name to extract first name and last name"""
                if fullname.startswith('นางสาว'):
                    name = fullname.replace('นางสาว', '').strip()
                elif fullname.startswith('นาย'):
                    name = fullname.replace('นาย', '').strip()
                else:
                    name = fullname.strip()
                
                parts = name.split()
                if len(parts) >= 2:
                    return parts[0], ' '.join(parts[1:])
                elif len(parts) == 1:
                    return parts[0], ''
                else:
                    return name, ''
            
            created_users = 0
            updated_users = 0
            added_enrollments = 0
            skipped_enrollments = 0
            
            for student_data in students_data:
                student_id = student_data['student_id']
                fullname = student_data['fullname']
                
                # Parse Thai name
                first_name, last_name = parse_thai_name(fullname)
                
                # Generate username from student_id (remove hyphen)
                username = student_id.replace('-', '')
                
                # Check if user already exists
                user, user_created = User.objects.get_or_create(
                    username=username,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'role': 'student',
                        'is_staff': False,
                        'email': f'{username}@student.example.com'
                    }
                )
                
                if user_created:
                    user.set_password('changeme123')
                    user.save()
                    created_users += 1
                else:
                    # Update existing user
                    user.first_name = first_name
                    user.last_name = last_name
                    user.role = 'student'
                    user.save()
                    updated_users += 1
                
                # Create or update profile
                profile, _ = UserProfile.objects.get_or_create(
                    user=user,
                    defaults={
                        'student_id': student_id
                    }
                )
                
                if profile.student_id != student_id:
                    profile.student_id = student_id
                    profile.save()
                
                # Add enrollment to bis3r1
                enrollment, enrollment_created = Enrollment.objects.get_or_create(
                    student=user,
                    section=section,
                    defaults={'status': 'enrolled'}
                )
                
                if enrollment_created:
                    added_enrollments += 1
                else:
                    skipped_enrollments += 1
            
            messages.success(
                request,
                f'เพิ่มนักเรียนเข้าในกลุ่มเรียน {section.course.course_code} - กลุ่ม {section.section_number} สำเร็จ!<br>'
                f'สร้างผู้ใช้ใหม่: {created_users} คน, อัปเดต: {updated_users} คน<br>'
                f'เพิ่มการลงทะเบียน: {added_enrollments} คน, ข้าม (ลงทะเบียนแล้ว): {skipped_enrollments} คน'
            )
            return redirect('academic:section_detail', section_id=section.id)
            
        except Exception as e:
            messages.error(request, f'เกิดข้อผิดพลาด: {str(e)}')
            return redirect('academic:section_list')
    
    # GET request - redirect to section_list
    return redirect('academic:section_list')

